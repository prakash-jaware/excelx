from flask import Flask, request, render_template, send_from_directory
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from datetime import datetime
import os
 
app = Flask(__name__)
 
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files.get("excel_file")
        username = request.form.get("YourUsername")
        if not file or not username:
            return "Please upload a file and enter username.", 400
 
        try:
            xl = pd.ExcelFile(file)
            df_list = xl.parse("Client_List")
            df_data = xl.parse("Data")
        except Exception as e:
            return f"❌ Error reading Excel file: {e}", 500
 
        base_path = Path(
            "C:/Users/YourUsername/OneDrive - Shree Plan Your Journey Pvt Ltd/Documents/SPYJ/RECON-24-25/2024-2025 Working/Online portal Sales & Refund/"
        )
 
        count = 0
        for _, row in df_list.iterrows():
            try:
                main_folder = str(row.iloc[0]).strip()
                sub_folder = str(row.iloc[1]).strip()
                date_value = row.iloc[2]
            except:
                continue
 
            if not main_folder or not sub_folder:
                continue
 
            try:
                month_start = pd.to_datetime(date_value).replace(day=1)
                month_end = pd.to_datetime(date_value).replace(day=28)  # You can use calendar.monthrange
                month_name = pd.to_datetime(date_value).strftime('%B')
            except:
                month_start = month_end = pd.Timestamp.now()
                month_name = "Unknown"
 
            folder_path = base_path / main_folder / month_name / sub_folder
            folder_path.mkdir(parents=True, exist_ok=True)
 
            client_name = sub_folder
            clean_name = ''.join(c for c in client_name if c not in r'[]:*?/\\')[:31] or "Sheet"
            df_client = df_data[df_data.iloc[:, 0].astype(str).str.strip() == client_name]
            if df_client.empty:
                continue
 
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            max_col = df_data.shape[1] + 1  # +1 for Sr No column
 
            # Styles
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_fill = PatternFill("solid", fgColor="FFFFCC")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
 
            # Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws.cell(row=1, column=1, value=f"{client_name.upper()}").font = Font(size=14, bold=True)
            ws.cell(row=1, column=1).alignment = center_align
 
            # Statement Range
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
            cell = ws.cell(row=2, column=1)
            cell.value = f"STATEMENT ON {month_start.strftime('%d/%m/%Y')} TO {month_end.strftime('%d/%m/%Y')}"
            cell.font = Font(size=12, bold=True)
            cell.alignment = center_align
 
            # Column Headers
            headers = ["Sr No"] + df_data.columns.tolist()
            ws.append(headers)
            for cell in ws[3]:
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill
 
            # Data Rows
            for idx, record in enumerate(df_client.itertuples(index=False), start=1):
                row = [idx] + list(record)
                ws.append(row)
                for col_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=idx + 3, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = center_align
 
            # TOTAL row
            total_row_idx = ws.max_row + 1
            total_label_col = len(headers) - 1
            total_value_col = len(headers)
            total_col_letter = ws.cell(row=3, column=total_value_col).column_letter
 
            ws.cell(row=total_row_idx, column=total_label_col, value="TOTAL")
            ws.cell(row=total_row_idx, column=total_label_col).font = bold_font
            ws.cell(row=total_row_idx, column=total_label_col).alignment = center_align
            ws.cell(row=total_row_idx, column=total_label_col).fill = yellow_fill
 
            ws.cell(row=total_row_idx, column=total_value_col,
                    value=f"=SUM({total_col_letter}4:{total_col_letter}{total_row_idx - 1})")
            ws.cell(row=total_row_idx, column=total_value_col).font = bold_font
            ws.cell(row=total_row_idx, column=total_value_col).alignment = center_align
            ws.cell(row=total_row_idx, column=total_value_col).fill = yellow_fill
 
            # GRAND TOTAL row
            grand_row_idx = total_row_idx + 1
            ws.merge_cells(start_row=grand_row_idx, start_column=1, end_row=grand_row_idx, end_column=total_label_col)
            ws.cell(row=grand_row_idx, column=1, value="GRAND TOTAL")
            ws.cell(row=grand_row_idx, column=1).font = Font(bold=True, size=12)
            ws.cell(row=grand_row_idx, column=1).alignment = center_align
            ws.cell(row=grand_row_idx, column=1).fill = yellow_fill
 
            ws.cell(row=grand_row_idx, column=total_value_col,
                    value=f"=SUM({total_col_letter}4:{total_col_letter}{total_row_idx - 1})")
            ws.cell(row=grand_row_idx, column=total_value_col).font = Font(bold=True, size=12)
            ws.cell(row=grand_row_idx, column=total_value_col).alignment = center_align
            ws.cell(row=grand_row_idx, column=total_value_col).fill = yellow_fill
 
            # Save
            output_file = folder_path / f"{clean_name}.xlsx"
            wb.save(output_file)
            count += 1
 
        return f"✅ {count} file(s) created successfully in '{base_path}'"
 
    return render_template("upload.html")
 
@app.route("/output/<path:filename>")
def download_file(filename):
    return send_from_directory(base_path, filename, as_attachment=True)
 

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
